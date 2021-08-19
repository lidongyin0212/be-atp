# -*- coding: utf-8 -*-
import openpyxl
from InterfaceTestManage.utils.sql_manager import *
from public.mysqldb import MySQLHelper
import json
from decimal import *


# 调试用连接数据库
# from public.db_operate.mysql_operate_db import MySQLHelper
# conn = MySQLHelper('10.8.72.59', 3306, 'root', 'a123456', 'test_zl')

def convert_export_data(ret, params, flag=False):
    field = ['id', 'name', 'rule', 'dataFormat'] if flag else ['name', 'rule', 'dataFormat']
    sql = SELECT_INTERFACE_EXTRACT_LIST if flag else SELECT_INTERFACE_EXTRACT_LIST2
    data_list = []
    for info in ret:
        count = 0
        data_dict = {}
        for item1 in info:
            if item1 is None:
                item1 = ''
            if params[count] == "extract_list" and item1:
                extract_list = str(item1) if item1 else ''
                sql1 = sql % "(" + extract_list + ")"
                result = MySQLHelper().get_all(sql1)
                item1 = json.dumps(convert_export(result, field, True))
            if params[count] == "state":
                item1 ="启用" if item1 == 1 else "禁用"
            if params[count] == "case_type":
                if item1 == 1:
                    item1 = "正常用例"
                if item1 == 2:
                    item1 = "前置用例"
                if item1 == 3:
                    item1 = "后置用例"
            data_dict[params[count]] = item1
            count += 1
        data_list.append(data_dict)
    return data_list


def convert_export(ret, params, flag=False):
    data_list = []
    for info in ret:
        count = 0
        data_dict = {}
        for item1 in info:
            if flag and isinstance(item1, str):
                if params[count] != 'except_list':
                    item1 = item1.replace("\\\\\\\\", "\\\\")
            data_dict[params[count]] = item1
            count += 1
        data_list.append(data_dict)
    return data_list


def exprot_case(subsystem_id):
    sql = SELECT_SUBSYSTEM_CASE % (subsystem_id)
    headers = {
        "case_name":"用例名称","req_path":"请求路径","req_method":"请求方法","req_headers":"请求头",
        "req_param":"请求参数","req_body":"请求体参数","req_file":"请求文件","extract_list":"提取变量",
        "except_list":"期望结果","run_time":"执行次数","wait_time":"等待时间",
        "project_name":"项目名称","subsystem_name":"子系统名称",
        "interface_name":"接口名称","inte_path":"接口路径","module_name":"模块名称","case_type":"类型",
        "username":"操作人","state":"状态"}
    param_sql = ['case_name','req_path','req_method','req_headers','req_param','req_body','req_file',
                 'extract_list','except_list','run_time','wait_time','project_name',
                 'subsystem_name','interface_name','inte_path','module_name','case_type','username','state']

    many_data = convert_export_data(MySQLHelper().get_all(sql), param_sql)

    return many_data, headers


def export_interface(subsystem_id):
    sql = SELECT_SUBSYSTEM_INTERFACE % (subsystem_id)
    headers = {
        "interface_name":"接口名称","req_path":"请求路径","req_method":"请求方法","req_headers":"请求头",
        "req_param":"请求参数","req_body":"请求体参数","req_file":"请求文件","extract_list":"提取变量",
        "except_list":"期望结果","project_name":"项目名称","subsystem_name":"子系统名称",
        "env_name":"环境名称","host":"主机地址","port":"端口号","username":"操作人","state":"状态","case_count":"用例数"
    }
    param_sql = ['interface_name','req_path','req_method','req_headers','req_param','req_body','req_file','extract_list',
                 'except_list','project_name','subsystem_name','env_name','host','port','username','state','case_count']
    interface_list = convert_export(MySQLHelper().get_all(sql),param_sql)
    for i in interface_list:
        if i["state"] == 1:
            i["state"] = "启用"
        else:
            i["state"] = "禁用"
        if i["case_count"]:
            i["case_count"] = int(i["case_count"])
    return interface_list, headers

def export_report(project_id):
    sql = EXPORT_REPORT % (project_id)
    headers = {
        "task_name":"任务名称","report_path":"报告路径","username":"操作人"
    }
    param_sql = ['task_name','report_path','username']
    report_list = convert_export(MySQLHelper().get_all(sql),param_sql)

    return report_list, headers

def export_task_report(task_id):
    sql = EXPORT_REPORT_TASK % (task_id)
    headers = {
        "task_name": "任务名称", "report_path": "报告路径", "username": "操作人"
    }
    param_sql = ['task_name', 'report_path', 'username']
    report_list = convert_export(MySQLHelper().get_all(sql), param_sql)
    return report_list, headers




# wb = openpyxl.Workbook()  # 新键工作簿
# w = wb.create_sheet(name, 0)  # 新键 sheet
# for i in range(len(total)):
#     w.cell(1, i + 1).value = total[i]
# for i in range(len(export_data)):
#     detail = list(export_data[i].values())
#     j = 1
#     for val in detail:
#         w.cell(i + 2, j).value = val
#         j += 1
# path = tempfile.mkdtemp(dir='.\static\export')  # 新键临时路径
# absolute_path = ("%s\%s.xlsx") % (path, name)
# exist_file = os.path.exists(absolute_path)
# if exist_file:
#     os.remove(absolute_path)
# wb.save(absolute_path)  # 保存指定路径
